# openpyxl
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
# worksheet: Worksheet = workbook.active  # type:ignore

sheet_name = 'minha planilha'
workbook.create_sheet(sheet_name)
worksheet: Worksheet = workbook[sheet_name]  # type:ignore

# remover uma planilha
workbook.remove(workbook['Sheet'])


# criando cabeçalhos
worksheet.cell(1, 1, 'Nome')  # linha, coluna, valor
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

students = [
    # nome idade nota
    ['joao', 14, 5.5],
    ['maria', 13, 9.7],
    ['luiz', 15, 8.8],
    ['alberto', 16, 10],
]

# for i, student_row in enumerate(students, start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_column)

for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)
