# openpyxl
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook: Workbook = load_workbook(WORKBOOK_PATH)

sheet_name = 'minha planilha'

worksheet: Worksheet = workbook[sheet_name]  # type:ignore

row: tuple[Cell]
for row in worksheet.iter_rows(min_row=2):
    for col in row:
        print(col.value, end='\t')

        if col.value == "maria":
            worksheet.cell(col.row, 2, 23)

    print()

# worksheet['B3'].value = 777

workbook.save(WORKBOOK_PATH)
